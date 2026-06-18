"""
coupang_bulk_20260618_124748_fixed.xlsm 에
 - 상세 설명  (R2 detail_combined URL 재구성)
 - 출고리드타임 (queue_state 또는 기본값 3)
 - 판매시작일  (오늘 날짜)
 를 채워넣는 패치 스크립트
"""
import json, re, datetime
from pathlib import Path
import openpyxl
import urllib.request

SRC_XLSX = Path.home() / "Downloads" / "coupang_bulk_20260618_124748.xlsx"
FIXED    = Path.home() / "Downloads" / "coupang_bulk_20260618_124748_fixed.xlsm"
R2_BASE  = "https://pub-52f3ccc0b1874a4dbca6ac2b8b860d49.r2.dev"

TODAY = datetime.datetime(2026, 6, 18, 0, 0, 0)

# ── queue_state.json 로드 (URL -> lead_time) ────────────────────────
qs_path = Path(__file__).parent / "data" / "queue_state.json"
url_to_lt = {}
if qs_path.exists():
    for q in json.loads(qs_path.read_text(encoding="utf-8")):
        lt = q.get("lead_time") or 3
        url_to_lt[q.get("url", "")] = int(lt)

def get_lead_time(naver_url):
    return url_to_lt.get(naver_url, 3)

# ── 네이버 URL에서 product_id 추출 ────────────────────────────────────
def extract_pid(naver_url):
    m = re.search(r'/products/(\d+)', str(naver_url or ""))
    return m.group(1) if m else ""

# ── R2 URL 존재 여부 확인 ─────────────────────────────────────────────
_url_cache = {}
def r2_exists(url):
    if url in _url_cache:
        return _url_cache[url]
    try:
        req = urllib.request.Request(url, method="HEAD")
        with urllib.request.urlopen(req, timeout=5) as r:
            ok = r.status == 200
    except Exception:
        ok = False
    _url_cache[url] = ok
    return ok

def get_detail_url(pid):
    if not pid:
        return ""
    combined = f"{R2_BASE}/{pid}_detail_combined.png"
    if r2_exists(combined):
        return combined
    # fallback: _detail_1.png, _detail.png
    for suffix in (f"{pid}_detail_1.png", f"{pid}_detail.png",
                   f"{pid}_1.jpg", f"{pid}_detail_img.png"):
        url = f"{R2_BASE}/{suffix}"
        if r2_exists(url):
            return url
    return ""

# ── 원본 xlsx에서 네이버URL 컬럼 읽기 ────────────────────────────────
wb_src = openpyxl.load_workbook(str(SRC_XLSX), read_only=True)
ws_src = wb_src.active
# 헤더 찾기
SRC_HDR = 2
src_cols = {}
for c in range(1, ws_src.max_column + 1):
    v = str(ws_src.cell(SRC_HDR, c).value or "").strip()
    if v:
        src_cols[v] = c

naver_col = src_cols.get("네이버원본URL")
SRC_DATA_START = SRC_HDR + 3
# row_idx -> naver_url
row_to_url = {}
for r in range(SRC_DATA_START, ws_src.max_row + 1):
    if naver_col:
        row_to_url[r - SRC_DATA_START] = ws_src.cell(r, naver_col).value or ""
print(f"[원본] {len(row_to_url)}행, 네이버URL 컬럼: {naver_col}")

# ── 변환 xlsm 로드 ─────────────────────────────────────────────────────
wb = openpyxl.load_workbook(str(FIXED), keep_vba=True)
ws = wb.active

# 헤더 찾기
HDR_R = 2
tmpl_cols = {}
for c in range(1, ws.max_column + 1):
    v = str(ws.cell(HDR_R, c).value or "").strip()
    if v:
        tmpl_cols[v] = c

DATA_START = HDR_R + 3

col_detail  = tmpl_cols.get("상세 설명")
col_lt      = tmpl_cols.get("출고리드타임")
col_start   = tmpl_cols.get("판매시작일")

print(f"[템플릿] 상세설명={col_detail}, 출고리드타임={col_lt}, 판매시작일={col_start}")

# ── 행별 패치 ─────────────────────────────────────────────────────────
patched = 0
for r in range(DATA_START, ws.max_row + 1):
    # 이 행에 데이터가 있는지 확인 (등록상품명 컬럼)
    name_col = tmpl_cols.get("등록상품명", 2)
    if not ws.cell(r, name_col).value:
        continue

    row_idx   = r - DATA_START
    naver_url = row_to_url.get(row_idx, "")
    pid       = extract_pid(naver_url)
    lead_time = get_lead_time(naver_url)

    # 상세 설명
    if col_detail and not ws.cell(r, col_detail).value:
        detail_url = get_detail_url(pid)
        if detail_url:
            ws.cell(r, col_detail).value = detail_url
            print(f"  {r}행: detail={detail_url[-40:]}")
        else:
            print(f"  {r}행: {pid} 상세이미지 R2에 없음")

    # 출고리드타임
    if col_lt:
        ws.cell(r, col_lt).value = lead_time

    # 판매시작일
    if col_start:
        ws.cell(r, col_start).value = TODAY

    patched += 1

wb.save(str(FIXED))
print(f"\n패치 완료: {patched}행 / {FIXED}")
