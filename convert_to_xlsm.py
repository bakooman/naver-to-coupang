"""
coupang_bulk_20260618_124748.xlsx -> coupang_bulk_20260618_124748_fixed.xlsm
표준모드 xlsx 컬럼명 -> Wing 템플릿 xlsm 컬럼명 명시 매핑 후 이식
"""
from pathlib import Path
import openpyxl

SRC  = Path.home() / "Downloads" / "coupang_bulk_20260618_124748.xlsx"
TMPL = Path(__file__).parent / "data" / "templates" / "wing_template.xlsm"
OUT  = Path.home() / "Downloads" / "coupang_bulk_20260618_124748_fixed.xlsm"

# 표준모드 컬럼명 -> Wing 템플릿 컬럼명 명시 매핑
EXPLICIT = {
    "카테고리아이디":   "카테고리",
    "등록상품명":       "등록상품명",
    "브랜드":           "브랜드",
    "제조사":           "제조사",
    "검색어":           "검색어",
    "판매가격":         "판매가격",
    "할인출기준가":     "할인율기준가",
    "재고수량":         "재고수량",
    "옵션유형1":        "옵션유형1",
    "옵션값1":          "옵션값1",
    "대표(옵션)이미지": "대표(옵션)이미지",
    "추가이미지1":      "추가이미지1",
    "추가이미지2":      "추가이미지2",
    "추가이미지3":      "추가이미지3",
    "네이버원본URL":    None,   # Wing 템플릿에 없음 → 건너뜀
}

def find_header_row(ws):
    for r in range(1, 10):
        vals = [str(ws.cell(r, c).value or "") for c in range(1, 30)]
        filled = [v for v in vals if v.strip() and len(v.strip()) <= 25]
        if len(filled) >= 5:
            return r
    return 2

# ── 소스 xlsx 읽기 ────────────────────────────────────────────────────
wb_src = openpyxl.load_workbook(str(SRC))
ws_src = wb_src.active
src_hdr = find_header_row(ws_src)
src_data_start = src_hdr + 3

src_col_idx = {}
for c in range(1, ws_src.max_column + 1):
    v = str(ws_src.cell(src_hdr, c).value or "").strip()
    if v:
        src_col_idx[v] = c

data_rows = []
for r in range(src_data_start, ws_src.max_row + 1):
    row = [ws_src.cell(r, c).value for c in range(1, ws_src.max_column + 1)]
    if any(v is not None for v in row):
        data_rows.append(row)

print(f"[소스] {len(data_rows)}행, 컬럼: {list(src_col_idx.keys())}")

# ── 템플릿 xlsm 로드 ──────────────────────────────────────────────────
wb_tmpl = openpyxl.load_workbook(str(TMPL), keep_vba=True)
ws_tmpl = wb_tmpl.active
tmpl_hdr = find_header_row(ws_tmpl)
tmpl_data_start = tmpl_hdr + 3

tmpl_col_idx = {}
for c in range(1, ws_tmpl.max_column + 1):
    v = str(ws_tmpl.cell(tmpl_hdr, c).value or "").strip()
    # 구매옵션/검색옵션 모두 "옵션유형1" 등 동일명 사용 → 첫 번째(구매옵션) 우선
    if v and v not in tmpl_col_idx:
        tmpl_col_idx[v] = c

# ── 컬럼 매핑 구성 ───────────────────────────────────────────────────
# src_col_num -> tmpl_col_num
col_map = {}
matched = []
skipped = []
for src_name, src_c in src_col_idx.items():
    tmpl_name = EXPLICIT.get(src_name, src_name)   # 명시 매핑, 없으면 동일명
    if tmpl_name is None:
        skipped.append(src_name)
        continue
    if tmpl_name in tmpl_col_idx:
        col_map[src_c] = tmpl_col_idx[tmpl_name]
        matched.append(f"{src_name}->{tmpl_name}")
    else:
        skipped.append(f"{src_name}(매핑:{tmpl_name} 없음)")

print(f"[매핑] 성공 {len(matched)}개: {matched}")
print(f"[스킵] {skipped}")

# ── 기존 데이터 행 삭제 ───────────────────────────────────────────────
for r in range(tmpl_data_start, ws_tmpl.max_row + 1):
    for c in range(1, ws_tmpl.max_column + 1):
        ws_tmpl.cell(r, c).value = None

# ── 데이터 이식 ──────────────────────────────────────────────────────
for row_idx, row_vals in enumerate(data_rows):
    tmpl_r = tmpl_data_start + row_idx
    for src_c, tmpl_c in col_map.items():
        val = row_vals[src_c - 1] if src_c - 1 < len(row_vals) else None
        ws_tmpl.cell(tmpl_r, tmpl_c).value = val

wb_tmpl.save(str(OUT))

# ── 결과 검증 ────────────────────────────────────────────────────────
wb_chk = openpyxl.load_workbook(str(OUT), keep_vba=True)
ws_chk = wb_chk.active
print(f"\n[완료] {OUT}")
print(f"검증 ({tmpl_data_start}~{tmpl_data_start+2}행):")
key_cols = ["카테고리", "등록상품명", "브랜드", "판매가격", "옵션유형1", "옵션값1", "대표(옵션)이미지"]
for r in range(tmpl_data_start, min(tmpl_data_start + 3, ws_chk.max_row + 1)):
    row_data = {k: ws_chk.cell(r, tmpl_col_idx[k]).value for k in key_cols if k in tmpl_col_idx}
    print(f"  {r}행: {row_data}")
