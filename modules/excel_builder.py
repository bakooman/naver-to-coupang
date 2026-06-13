"""
쿠팡 Wing 일괄등록 Excel 생성기 (v2 — 실제 템플릿 구조 반영)

Wing 제공 .xlsm 템플릿 행 구조:
  1행: 그룹 헤더 (기본정보 / 구매옵션 / 검색옵션 / 고시정보 / 이미지 / ...)
  2행: 컬럼명  (카테고리, 등록상품명, 브랜드, 판매가격, 대표(옵션)이미지, ...)
  3행: 가이드  (필수/선택 표시 및 입력 안내 — 긴 텍스트)
  4행~: 데이터

주요 수정사항 (v1 → v2):
  - _find_header_row: 셀 길이 ≤ 20 필터로 가이드 행(긴 텍스트) 제외 → 2행 정확히 탐지
  - data_start = header_row + 2  (3행 가이드 건너뜀)
  - 컬럼명 수정: 등록상품명 / 판매가격 / 할인출기준가 / 대표(옵션)이미지
  - 옵션유형·옵션값: 구매옵션 첫번째만 매핑 (수량 / N개)
  - 추가이미지: 발견 순서대로 _extra_img_0 … _extra_img_N 매핑
  - 상품고시정보 보고N: 전부 "상세페이지 참조" 로 일괄 기입
  - 템플릿 확장자 보존 (.xlsm → .xlsm 저장, VBA 유지)
  - 필수 항목만 채움 (선택 항목 생략)
"""
from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── 데이터 모델 ────────────────────────────────────────────────────

@dataclass
class Bundle:
    qty:            int    # 묶음 수량 (1, 2, 3 ...)
    sale_price:     int    # 판매가격
    original_price: int    # 할인출기준가 (정상가)
    image_url:      str = ""   # R2 URL
    option_label:   str = ""   # 네이버 옵션명 (옵션별 가격 분리 시 사용)


@dataclass
class BulkItem:
    naver_url:          str
    product_name:       str
    brand:              str
    category_id:        str = ""
    bundles:            list[Bundle] = field(default_factory=list)
    main_image_url:     str = ""
    extra_image_urls:   list[str] = field(default_factory=list)
    origin:             str = "수입산"
    manufacturer:       str = ""
    model_number:       str = ""
    tags:               list[str] = field(default_factory=list)
    stock:              int = 500
    error:              str = ""
    detail_description: str = ""   # 상세 설명 HTML (DF 컬럼, 필수)
    extra_options:      list[tuple[str, str]] = field(default_factory=list)
    gosisi_cat:         str = ""   # 상품고시정보 카테고리 (비워두면 _GOSISI_CAT_MAP 자동 탐색)
    draft:              bool = False  # True → 판매시작일=None → Wing 임시저장 (상세페이지 직접 수정용)
    gtin:               str = ""   # 바코드(GTIN) 8~14자리 숫자 — GS1 Korea 역조회 or 수동입력
    qty_unit:           str = "개" # 수량 단위: "개" (일반) / "세트" (교환세트)
    lead_time:          int = 3   # 출고리드타임 (국내 3일 / 해외 10일)
    qty_as_option:      bool = True  # False → 수량 슬롯 생략, extra_options[0]부터 슬롯0 사용
    qty_option_type:    str  = "수량" # 옵션유형 이름: "수량"(기본) or "총 수량"(컵라면 등)
    # extra_options 예시:
    #   [("엔진오일 SAE점도", "5w30"), ("개당 용량", "1L")]  ← 엔진오일
    #   [("중량", "500g")]                                   ← 식품
    #   [("색상", "블랙"), ("사이즈", "55인치")]              ← 가전
    # 슬롯 0 = 수량(자동), 슬롯 1~2 = extra_options 순서대로 삽입
    #
    # gosisi_cat 권장값:
    #   "화장품"                         ← 뷰티 전체
    #   "가공식품"                       ← 식품류
    #   "건강기능식품"                   ← 건강기능식품
    #   "자동차용품"                     ← 엔진오일
    #   "가정용 전기제품(냉장고/세탁기 등)" ← 소형가전/주방가전
    #   "기타 재화"                      ← 기본값


# ── 상품고시정보 카테고리 매핑 ─────────────────────────────────────
# 쿠팡 카테고리 ID → 상품고시정보 카테고리 드롭다운 값
_GOSISI_CAT_MAP: dict[str, str] = {
    "78889": "자동차용품",
    "78897": "자동차용품",
    "78893": "자동차용품",
    "78903": "자동차용품",  # 부동액/냉각수
    "78894": "자동차용품",  # 브레이크오일
}
_GOSISI_CAT_DEFAULT = "기타 재화"
_GOSISI_ITEM_VALUE  = "상세페이지 참조"  # 각 보고 항목 공통 기입값


# ── ExcelBuilder ───────────────────────────────────────────────────

class ExcelBuilder:
    """BulkItem 목록 → 쿠팡 Wing 일괄등록 Excel 파일 생성."""

    # ── 표준 컬럼 (템플릿 없을 때 fallback) ──────────────────────
    _STD_COLUMNS: list[tuple[str, int]] = [
        ("카테고리아이디",    14),
        ("등록상품명",        40),
        ("브랜드",            16),
        ("제조사",            16),
        ("검색어",            30),
        ("판매가격",          12),
        ("할인출기준가",      12),
        ("재고수량",          10),
        ("옵션유형1",         14),
        ("옵션값1",           14),
        ("대표(옵션)이미지",  60),
        ("추가이미지1",       60),
        ("추가이미지2",       60),
        ("추가이미지3",       60),
        ("네이버원본URL",     50),
    ]

    # ── 헤더 행 감지 키워드 (셀 길이 ≤ 20 && 키워드 포함 시 카운트) ──
    _HEADER_KEYWORDS: frozenset[str] = frozenset({
        "카테고리", "등록상품명", "브랜드", "제조사",
        "검색어",   "판매가격",   "재고수량",
        "대표(옵션)이미지", "추가이미지", "상품고시정보",
        "옵션유형", "옵션값",     "모델번호",  "출고리드타임",
        "할인율기준가",
    })

    # ── 단순 컬럼 매핑 (헤더 substring → 내부 키) ────────────────
    # 특수 처리 제외: 옵션유형, 옵션값, 추가이미지, 상품고시정보값N
    # ⚠️ 순서 중요: 더 구체적인 키워드를 먼저 배치해야 함
    #   "상품고시정보 카테고리"는 "카테고리" 부분문자열을 포함하므로
    #   반드시 "카테고리" 매핑보다 앞에 위치해야 올바르게 분리됨.
    _TEMPLATE_MAP: dict[str, str] = {
        # 판매시작일: 오늘 날짜 기입 → 즉시 판매 시작 (공란이면 임시저장됨)
        # 판매종료일: 공란 유지 → 종료일 없음 (무기한 판매)
        "판매시작일":            "_sale_start",
        "판매종료일":            "_sale_end",
        # ⚠️ "상품고시정보 카테고리"를 "카테고리" 앞에 배치 (substring 충돌 방지)
        "상품고시정보 카테고리": "_gosisi_cat",
        "카테고리":              "category_id",
        "등록상품명":            "product_name",
        "브랜드":                "brand",
        "제조사":                "manufacturer",
        "모델번호":              "model_number",
        # "검색어" 컬럼은 매핑하지 않음 — 형식 오류로 등록 실패 유발 (선택 항목)
        "판매가격":              "_sale_price",       # 판매가 (×1.37)
        "할인율기준가":          "_original_price",   # 필수항목 — 판매가와 동일값 기입 (할인배지 미노출)
        # 자동가격조정(최저가/설정가격)은 엑셀 템플릿에 컬럼 없음 → Wing UI에서만 설정 가능
        "재고수량":              "stock",
        "출고리드타임":          "_lead_time",
        "바코드":                "gtin",
        "대표(옵션)이미지":      "_main_img",
        "상세 설명":             "_detail_desc",      # DF 컬럼 (필수 — 빈값 시 등록 실패)
        "초도반품배송비":        "_init_return_fee",  # 0으로 고정 — 카테고리별 상한 초과 방지
        "반품배송비":            "_return_fee",       # 0으로 고정 — 카테고리별 상한 초과 방지
        "해외구매대행":          "_overseas",         # Y=구매대행(개인통관고유부호 고객 입력 필수), N=일반
    }

    def __init__(
        self,
        template_path: Optional[str | Path] = None,
        output_dir: str | Path = "data/output",
        category_id: str = "",
    ):
        if not OPENPYXL_OK:
            raise RuntimeError("openpyxl 미설치: pip install openpyxl")
        self._template    = Path(template_path) if template_path else None
        self._output_dir  = Path(output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)
        self._category_id = category_id

    # ── Public ─────────────────────────────────────────────────────

    def build(
        self,
        items: list[BulkItem],
        filename: str | None = None,
        lead_time: int = 3,
    ) -> Path:
        self._lead_time = lead_time   # 출고리드타임: 국내 3일 / 해외 10일
        if self._template and self._template.exists():
            wb  = self._fill_template(items)
            ext = self._template.suffix.lower()   # .xlsm 유지
        else:
            wb  = self._create_standard(items)
            ext = ".xlsx"

        ts    = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = filename or f"coupang_bulk_{ts}{ext}"
        out   = self._output_dir / fname
        wb.save(str(out))
        print(f"[Excel] 파일 저장: {out}")
        return out

    # ── 템플릿 기반 ───────────────────────────────────────────────

    def _fill_template(self, items: list[BulkItem]) -> "openpyxl.Workbook":
        """사용자 제공 Wing 템플릿 로드 → 데이터 삽입."""
        wb = load_workbook(str(self._template), keep_vba=True)
        ws = wb.active

        header_row = self._find_header_row(ws)
        if header_row is None:
            print("[Excel] 헤더 행 감지 실패 → 표준 포맷 대체")
            return self._create_standard(items)

        print(f"[Excel] 헤더 행 감지: {header_row}행")
        col_map = self._map_template_columns(ws, header_row)
        print(f"[Excel] 매핑된 컬럼 ({len(col_map)}개): {sorted(col_map.keys())}")

        # 데이터 시작 행: 헤더+3
        # 1행=그룹헤더, 2행=컬럼명(헤더), 3행=가이드설명, 4행=추가가이드 → 5행부터 데이터
        data_start = header_row + 3
        if ws.max_row >= data_start:
            ws.delete_rows(data_start, ws.max_row - data_start + 1)

        row_idx = data_start
        for item in items:
            if item.error:
                continue
            for bundle in sorted(item.bundles, key=lambda b: b.qty):
                self._write_row_template(ws, row_idx, col_map, item, bundle)
                row_idx += 1

        return wb

    def _find_header_row(self, ws) -> Optional[int]:
        """
        알려진 컬럼명 키워드가 가장 많이 등장하는 행을 헤더로 판별.
        셀 값 길이 > 20 인 셀은 가이드 설명으로 간주해 제외.
        """
        best_row, best_count = None, 0
        for row in ws.iter_rows(min_row=1, max_row=10):
            count = 0
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                val = cell.value.strip()
                if len(val) > 20:          # 가이드 설명 행(긴 텍스트) 제외
                    continue
                for kw in self._HEADER_KEYWORDS:
                    if kw in val:
                        count += 1
                        break
            if count > best_count:
                best_count, best_row = count, row[0].row

        return best_row if best_count >= 3 else None

    def _map_template_columns(self, ws, header_row: int) -> dict[str, int]:
        """
        헤더 행 순회 → 내부 키 → 열 인덱스(1-based) 매핑.

        특수 처리:
          옵션유형    → 발견 순서대로 → _option_type_0, _option_type_1, _option_type_2 ...
          옵션값      → 발견 순서대로 → _option_val_0,  _option_val_1,  _option_val_2  ...
                        (수량=0, SAE점도=1, 개당용량=2)
          추가이미지  → 발견 순서대로 → _extra_img_0, _1, ...
          상품고시정보값N → 발견 순서대로 → _gosisi_val_0, _1, ...
        """
        col_map: dict[str, int] = {}
        extra_img_idx   = 0
        gosisi_val_idx  = 0
        option_type_idx = 0   # 옵션유형 슬롯 (0=수량, 1=SAE점도, 2=개당용량)
        option_val_idx  = 0   # 옵션값   슬롯 (0=N개,   1=0w20,    2=1L)

        for cell in ws[header_row]:
            if not cell.value or not isinstance(cell.value, str):
                continue
            val = cell.value.strip()

            # 옵션유형 — 발견 순서대로 인덱스 부여 (최대 3슬롯 사용)
            if "옵션유형" in val and option_type_idx < 3:
                col_map[f"_option_type_{option_type_idx}"] = cell.column
                option_type_idx += 1
                continue

            # 옵션값 — 발견 순서대로 인덱스 부여 (최대 3슬롯 사용)
            if "옵션값" in val and option_val_idx < 3:
                col_map[f"_option_val_{option_val_idx}"] = cell.column
                option_val_idx += 1
                continue

            # 추가이미지 (상태이미지 제외)
            if "추가이미지" in val and "상태" not in val:
                col_map[f"_extra_img_{extra_img_idx}"] = cell.column
                extra_img_idx += 1
                continue

            # 상품고시정보값N — 실제 컬럼명: "상품고시정보값1" ~ "상품고시정보값14"
            if "상품고시정보값" in val:
                col_map[f"_gosisi_val_{gosisi_val_idx}"] = cell.column
                gosisi_val_idx += 1
                continue

            # 단순 매핑 (_TEMPLATE_MAP)
            for kw, internal in self._TEMPLATE_MAP.items():
                if kw in val and internal not in col_map:
                    col_map[internal] = cell.column
                    break

        print(
            f"[Excel] 추가이미지 {extra_img_idx}열 | "
            f"고시정보 보고 {gosisi_val_idx}열 | "
            f"옵션유형슬롯 {option_type_idx}개 | "
            f"옵션값슬롯 {option_val_idx}개 | "
            f"카테고리={'O(col ' + str(col_map['category_id']) + ')' if 'category_id' in col_map else 'X(미감지!)'} | "
            f"고시정보카테고리={'O' if '_gosisi_cat' in col_map else 'X'}"
        )
        return col_map

    def _write_row_template(
        self, ws, row: int, col_map: dict, item: BulkItem, bundle: Bundle
    ):
        """하나의 번들(수량) 행을 필수 필드 위주로 기록."""
        def w(key: str, value):
            if key in col_map and value is not None and value != "":
                ws.cell(row=row, column=col_map[key], value=value)

        # Wing 카테고리 ID: 숫자 문자열 → int 변환 (Excel 셀 타입이 숫자여야 Wing 검증 통과)
        raw_cat = (item.category_id or self._category_id).strip()
        cat_id  = int(raw_cat) if raw_cat.isdigit() else raw_cat
        mfr     = item.manufacturer or item.brand

        # ── 기본 정보 (필수) ───────────────────────────────────────
        # Wing Excel 업로드는 항상 "임시저장" 상태로 생성됨 (즉시 판매중 불가)
        # 판매시작일 제어 방식:
        #   draft=False → 어제 날짜 → Wing "설정함 + 이미 지난 시작일" → 업로드 후 자동 활성화
        #   draft=True  → None(공란) → Wing "영구 임시저장" → 상세페이지 수정 후 수동 등록 필요
        # 판매종료일: 항상 None → 종료일 없음 (무기한)
        _yesterday = datetime.date.today() - datetime.timedelta(days=1)
        if "_sale_start" in col_map:
            ws.cell(
                row=row, column=col_map["_sale_start"],
                value=None if item.draft else _yesterday,
            )
        if "_sale_end" in col_map:
            ws.cell(row=row, column=col_map["_sale_end"], value=None)
        w("category_id",  cat_id)
        w("product_name", item.product_name)
        w("brand",        item.brand)
        w("manufacturer", mfr)
        w("model_number", item.model_number)

        # ── 가격 / 재고 / 출고리드타임 ────────────────────────────────
        # 판매가격       = ×1.37 (bundle.original_price)
        # 할인율기준가   = ×1.37 (bundle.original_price) ← 필수항목이라 판매가와 동일하게 기입
        #                  → 정상가=판매가이므로 할인배지 미노출 (0% 할인)
        # 최저가         = ×1.27 (bundle.sale_price)   — 자동가격조정 하한
        # 설정가격       = ×1.37 (bundle.original_price) — 자동가격조정 목표
        w("_sale_price",     bundle.original_price)  # 판매가 (×1.37)
        w("_original_price", bundle.original_price)  # 할인율기준가 = 판매가 동일 (필수항목, 할인배지 미노출)
        # 자동가격조정 최저가/설정가격: 템플릿에 컬럼 없음 → Wing UI에서 직접 설정
        w("stock",        item.stock)
        w("_lead_time",   item.lead_time)  # 출고리드타임: 국내 3일 / 해외 10일 (per-item)
        # 바코드(GTIN): 최소 수량 번들 행에만 기입
        # — 쿠팡은 묶음별 GTIN이 달라야 하므로 2개·3개 번들에는 실제 바코드가 없음
        # — 동일 GTIN을 여러 행에 쓰면 "중복된 바코드" 오류 발생
        _min_qty = min((b.qty for b in item.bundles), default=bundle.qty)
        if (bundle.qty == _min_qty
                and item.gtin and item.gtin.isdigit()
                and 8 <= len(item.gtin) <= 14):
            w("gtin", item.gtin)

        # 반품배송비: 0원으로 고정 — 카테고리별 상한(예: 29,160원) 초과 방지
        # 템플릿 기본값이 상한을 초과하면 Wing 업로드 실패 원인이 됨
        w("_init_return_fee", 0)
        w("_return_fee",      0)

        # 해외구매대행: 해외배송만 Y 기입, 국내는 빈칸 (Wing이 자동으로 N 처리)
        # 명시적으로 N을 쓰면 Wing이 "배송방법을 확인해 주시기 바랍니다" 오류 발생
        if item.lead_time == 10:
            w("_overseas", "Y")

        # ── 구매 옵션 ───────────────────────────────────────────────────
        if item.qty_as_option:
            # 슬롯 0: 수량 (기본) / 슬롯 1~2: extra_options
            w("_option_type_0", item.qty_option_type)   # "수량" 또는 "총 수량" 등
            w("_option_val_0",  f"{bundle.qty}{item.qty_unit}")
            for _slot, (_otype, _oval) in enumerate(item.extra_options, 1):
                if _slot >= 3:
                    break
                w(f"_option_type_{_slot}", _otype)
                w(f"_option_val_{_slot}",  _oval)
        else:
            # 수량 옵션 불허 카테고리 (예: 63726 커피그라인더 → 모델명/품번만 허용)
            # extra_options[0]부터 슬롯 0에 배치
            for _slot, (_otype, _oval) in enumerate(item.extra_options):
                if _slot >= 3:
                    break
                w(f"_option_type_{_slot}", _otype)
                w(f"_option_val_{_slot}",  _oval)

        # ── 대표 이미지 URL (수량별 번들 이미지) ─────────────────────
        main_img = bundle.image_url or item.main_image_url
        w("_main_img", main_img)
        # 추가이미지는 등록하지 않음 — 모두 1개 배지 이미지가 들어가는 문제 방지

        # ── 상세 설명 (필수) ───────────────────────────────────────
        # Wing 서버는 DF 컬럼 값을 JSON으로 파싱하는 경우가 있어
        # raw HTML(<img ...>)을 그대로 쓰면 '<' 문자에서 JsonParseException 발생.
        # → HTML에서 이미지 src URL만 추출해 첫 번째 URL을 대표값으로 기입.
        #   URL이 없으면 HTML 전체를 그대로 기입(기존 동작 유지).
        if item.detail_description:
            # detail_description은 이제 합성 이미지 단일 URL (또는 fallback URL).
            # HTML 태그 없이 순수 URL이므로 그대로 기입.
            # (이전: HTML <img src='...'> 형식 → Wing 일부 카테고리에서 JsonParseException 발생)
            w("_detail_desc", item.detail_description)

        # ── 검색어: 쿠팡이 행별 개별 컬럼 형식을 요구 → 형식 오류 방지를 위해 미입력 ─
        # (선택 항목이며, 잘못된 형식 입력 시 등록 전체 실패 원인이 됨)

        # ── 상품고시정보 (필수) ────────────────────────────────────
        # 우선순위: 1) item.gosisi_cat 직접 지정 → 2) ID 맵 자동 탐색 → 3) 기본값
        # _GOSISI_CAT_MAP 키는 문자열이므로 str(cat_id) 로 조회
        gosisi_cat = (
            item.gosisi_cat.strip()
            or _GOSISI_CAT_MAP.get(str(cat_id), _GOSISI_CAT_DEFAULT)
        )
        w("_gosisi_cat", gosisi_cat)
        # 상품고시정보값1~N: "상세페이지 참조" 일괄 기입 (최대 20개 커버)
        for i in range(20):
            key = f"_gosisi_val_{i}"
            if key in col_map:
                ws.cell(row=row, column=col_map[key], value=_GOSISI_ITEM_VALUE)

    # ── 표준 포맷 (템플릿 없을 때 fallback) ──────────────────────

    def _create_standard(self, items: list[BulkItem]) -> "openpyxl.Workbook":
        """Wing 템플릿 없을 때 최소 표준 포맷 생성."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "쿠팡일괄등록"

        ws.merge_cells("A1:O1")
        ws["A1"] = (
            "※ 표준 포맷 (템플릿 없음) — Wing 다운로드 .xlsm 을 "
            "data/templates/ 에 넣으면 정확한 컬럼에 자동 입력됩니다."
        )
        ws["A1"].font = Font(color="FF0000", bold=True)

        headers = [c[0] for c in self._STD_COLUMNS]
        widths  = [c[1] for c in self._STD_COLUMNS]
        fill_h  = PatternFill("solid", fgColor="1F4E79")
        font_h  = Font(color="FFFFFF", bold=True)
        align_h = Alignment(horizontal="center", vertical="center")

        for ci, (hdr, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=2, column=ci, value=hdr)
            cell.fill      = fill_h
            cell.font      = font_h
            cell.alignment = align_h
            ws.column_dimensions[get_column_letter(ci)].width = w

        row_idx = 3
        for item in items:
            if item.error:
                continue
            raw_cat = (item.category_id or self._category_id).strip()
            cat_id  = int(raw_cat) if raw_cat.isdigit() else raw_cat
            mfr     = item.manufacturer or item.brand
            for bundle in sorted(item.bundles, key=lambda b: b.qty):
                data: dict[str, object] = {
                    "카테고리아이디":    cat_id,
                    "등록상품명":        item.product_name,
                    "브랜드":            item.brand,
                    "제조사":            mfr,
                    "검색어":            ",".join(item.tags[:5]),
                    "판매가격":          bundle.sale_price,
                    "할인출기준가":      bundle.original_price,
                    "재고수량":          item.stock,
                    "옵션유형1":         "수량",
                    "옵션값1":           f"{bundle.qty}개",
                    "대표(옵션)이미지":  bundle.image_url or item.main_image_url,
                    "추가이미지1":       item.extra_image_urls[0] if len(item.extra_image_urls) > 0 else "",
                    "추가이미지2":       item.extra_image_urls[1] if len(item.extra_image_urls) > 1 else "",
                    "추가이미지3":       item.extra_image_urls[2] if len(item.extra_image_urls) > 2 else "",
                    "네이버원본URL":     item.naver_url,
                }
                for ci, hdr in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=ci, value=data.get(hdr, ""))
                row_idx += 1

        ws.freeze_panes = "A3"
        return wb
