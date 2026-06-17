"""
오류 엑셀 자동 수정 모듈

Wing 일괄등록 Excel을 파싱 → Gemini로 카테고리/브랜드 오매핑 감지 → 수정 Excel 재출력.
옵션 중복 제거 및 필수 옵션 타입 자동 채우기 포함.
"""
from __future__ import annotations

import datetime
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── 데이터 모델 ────────────────────────────────────────────────────

@dataclass
class FixResult:
    product_name:      str
    old_category_id:   str
    old_category_name: str
    new_category_id:   str
    new_category_name: str
    new_gosisi_cat:    str            = ""
    old_brand:         str            = ""
    new_brand:         str            = ""
    needs_fix:         bool           = False
    reason:            str            = ""
    brand_locked:      bool           = False
    row_indices:       list[int]      = field(default_factory=list)
    # 옵션 관련
    dup_rows:               list[int]       = field(default_factory=list)
    invalid_option_types:   list[str]       = field(default_factory=list)
    missing_required_opts:  list[str]       = field(default_factory=list)
    missing_opt_values:     dict[str, str]  = field(default_factory=dict)  # type→Gemini 추정값
    has_option_issues:      bool            = False


# ── 헤더 감지 (ExcelBuilder와 동일 로직) ──────────────────────────

_HEADER_KW = frozenset([
    "카테고리", "등록상품명", "브랜드", "판매가격",
    "옵션유형", "옵션값", "재고수량", "출고리드타임",
])


def _find_header_row(ws) -> Optional[int]:
    best_row, best_cnt = None, 0
    for row in ws.iter_rows(min_row=1, max_row=10):
        cnt = 0
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            val = cell.value.strip()
            if len(val) > 20:
                continue
            if any(kw in val for kw in _HEADER_KW):
                cnt += 1
        if cnt > best_cnt:
            best_cnt, best_row = cnt, row[0].row
    return best_row if best_cnt >= 3 else None


def _map_columns(ws, header_row: int) -> dict[str, int]:
    """헤더 행 → 내부 키 → 열 인덱스(1-based) 매핑."""
    col_map: dict[str, int] = {}
    ot_idx = ov_idx = 0
    for cell in ws[header_row]:
        if not isinstance(cell.value, str):
            continue
        v = cell.value.strip()
        if "상품고시정보 카테고리" in v:
            col_map["_gosisi_cat"] = cell.column
        elif "카테고리" in v and "상품고시정보" not in v and "category_id" not in col_map:
            col_map["category_id"] = cell.column
        elif "등록상품명" in v and "product_name" not in col_map:
            col_map["product_name"] = cell.column
        elif "브랜드" in v and "brand" not in col_map and "제조사" not in v:
            col_map["brand"] = cell.column
        elif "옵션유형" in v and ot_idx < 3:
            col_map[f"_option_type_{ot_idx}"] = cell.column
            ot_idx += 1
        elif "옵션값" in v and ov_idx < 3:
            col_map[f"_option_val_{ov_idx}"] = cell.column
            ov_idx += 1
    return col_map


# ── Excel 파싱 ────────────────────────────────────────────────────

def parse_excel(src_path: Path) -> tuple[dict, str, list[dict], int]:
    """
    Wing Excel 파싱.

    Returns:
        (col_map, src_ext, products, header_row)
        products: 상품명 기준으로 그룹핑된 dict 리스트
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl 미설치: pip install openpyxl")

    src_ext = src_path.suffix.lower() or ".xlsx"
    wb = load_workbook(str(src_path), keep_vba=(src_ext == ".xlsm"), data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    if not header_row:
        raise ValueError("헤더 행을 찾을 수 없습니다. Wing 일괄등록 Excel인지 확인하세요.")

    col_map = _map_columns(ws, header_row)
    if "product_name" not in col_map:
        raise ValueError("'등록상품명' 컬럼을 찾을 수 없습니다.")

    def _get(row_idx: int, key: str) -> str:
        col = col_map.get(key)
        if not col:
            return ""
        v = ws.cell(row=row_idx, column=col).value
        return str(v).strip() if v is not None else ""

    data_start = header_row + 3
    seen: dict[str, dict] = {}

    for r in range(data_start, ws.max_row + 1):
        pname = _get(r, "product_name")
        if not pname:
            continue

        row_opts: list[tuple[str, str]] = []
        for i in range(3):
            ot = _get(r, f"_option_type_{i}")
            ov = _get(r, f"_option_val_{i}")
            if ot:
                row_opts.append((ot, ov))

        if pname not in seen:
            seen[pname] = dict(
                product_name=pname,
                category_id=_get(r, "category_id"),
                brand=_get(r, "brand"),
                options=row_opts,
                all_options=[],
                options_by_row={},
                seen_combos=set(),
                dup_rows=[],
                row_indices=[],
            )

        prod = seen[pname]
        combo_key = tuple(row_opts)

        if combo_key in prod["seen_combos"]:
            prod["dup_rows"].append(r)
        else:
            prod["seen_combos"].add(combo_key)
            for pair in row_opts:
                prod["all_options"].append(pair)

        prod["options_by_row"][r] = row_opts
        prod["row_indices"].append(r)

    for p in seen.values():
        p.pop("seen_combos", None)

    wb.close()
    return col_map, src_ext, list(seen.values()), header_row


# ── 옵션 검증 ─────────────────────────────────────────────────────

def check_options(products: list[dict], cat_options_path: Path) -> list[dict]:
    """
    category_options.json 기준으로 각 상품 옵션 검증.

    추가 필드:
      - invalid_option_types: valid_options 외 사용된 옵션 타입
      - missing_required_opts: 한 번도 사용되지 않은 required_options 타입
      - has_option_issues: bool
    """
    try:
        with open(cat_options_path, encoding="utf-8") as f:
            cat_opts_data: dict = json.load(f)
    except Exception:
        cat_opts_data = {}

    for prod in products:
        cat_id   = prod.get("category_id", "")
        cat_info = cat_opts_data.get(str(cat_id), {})
        valid_opts    = set(cat_info.get("valid_options", []))
        required_opts = set(cat_info.get("required_options", []))

        used_types: set[str] = set()
        for r, row_opts in prod.get("options_by_row", {}).items():
            if r not in prod.get("dup_rows", []):
                for ot, _ in row_opts:
                    if ot:
                        used_types.add(ot)

        # 유효하지 않은 타입 (valid_options 있을 때만)
        invalid: list[str] = []
        if valid_opts:
            for t in sorted(used_types):
                if t not in valid_opts:
                    invalid.append(t)

        # 누락된 필수 옵션 (required_options 중 단 하나도 없는 타입만 개별 플래그)
        missing: list[str] = []
        if required_opts:
            for req in sorted(required_opts):
                if req not in used_types:
                    missing.append(req)

        prod["invalid_option_types"] = invalid
        prod["missing_required_opts"] = missing
        prod["has_option_issues"] = bool(
            prod.get("dup_rows") or invalid or missing
        )

    return products


# ── Gemini 검수 (브랜드/카테고리 + 누락 옵션 값 추정) ──────────────

def gemini_check(
    product: dict,
    api_key: str,
    model: str,
    missing_opts: list[str] | None = None,
) -> dict:
    """
    Gemini로 상품의 카테고리/브랜드 정합성 검수.
    missing_opts가 있으면 해당 옵션 타입의 대표 값도 추정.

    Returns:
        {needs_fix, category_keyword, brand, reason, option_values}
    """
    pname    = product["product_name"]
    cat_id   = product.get("category_id", "")
    cat_name = product.get("_cat_name", "")
    brand    = product.get("brand", "")
    options  = product.get("options", [])
    opts_str = ", ".join(f"{t}={v}" for t, v in options) if options else "없음"

    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)

        opt_ask = ""
        opt_json = ""
        if missing_opts:
            opt_ask = (
                f"\n⚠️ 아래 필수 옵션 타입이 비어 있습니다. 상품명을 근거로 대표 값을 추정하세요:\n"
                + "".join(f"  - {o}\n" for o in missing_opts)
            )
            opt_json = ', "option_values": {"옵션타입": "추정값", ...}'

        prompt = (
            "쿠팡 Wing 일괄등록 엑셀 상품 데이터를 검수합니다.\n\n"
            f"상품명: {pname}\n"
            f"카테고리 ID: {cat_id} / 카테고리명: {cat_name or '미확인'}\n"
            f"브랜드: {brand or '없음'}\n"
            f"옵션: {opts_str}\n"
            f"{opt_ask}\n"
            "⚠️ 필수 분류 규칙:\n"
            "1. 반려동물+건강보조 키워드 동시 → '강아지영양제'/'고양이영양제'. 인체용 건강기능식품 금지.\n"
            "2. 젤리/사탕/구미/캔디/스낵 → '식품간식'. 화장품·뷰티 금지.\n"
            "3. 카테고리명이 상품과 명백히 불일치할 때만 needs_fix=true.\n"
            "4. 브랜드: 쿠팡 공식 등록 브랜드명 그대로 반환. 확실치 않으면 원본 유지.\n\n"
            f"아래 JSON 형식으로만 답하세요:\n"
            '{"needs_fix": true또는false, '
            '"category_keyword": "올바른카테고리한국어키워드", '
            '"brand": "올바른브랜드명", '
            f'"reason": "이유15자이내"{opt_json}}}'
        )

        resp = genai.GenerativeModel(model).generate_content(prompt)
        raw  = (resp.text or "").strip()
        m    = re.search(r'\{[^{}]*\}', raw, re.DOTALL)
        if m:
            result = json.loads(m.group())
            if not isinstance(result.get("needs_fix"), bool):
                result["needs_fix"] = False
            result.setdefault("category_keyword", "")
            result.setdefault("brand", brand)
            result.setdefault("reason", "")
            result.setdefault("option_values", {})
            # option_values 유효성 검증
            if not isinstance(result["option_values"], dict):
                result["option_values"] = {}
            return result

    except json.JSONDecodeError:
        pass
    except Exception as e:
        print(f"[ExcelFixer] Gemini 오류 ({pname[:20]}): {e}")

    return {
        "needs_fix": False, "category_keyword": "", "brand": brand,
        "reason": "", "option_values": {},
    }


# ── 수정 적용 & 저장 ─────────────────────────────────────────────

def apply_and_save(
    src_path: Path,
    col_map: dict,
    fix_results: list[FixResult],
    output_dir: Path,
    src_ext: str = ".xlsx",
) -> Path:
    """
    원본 파일을 재로드하여 수정사항을 적용하고 저장.
    - 카테고리/브랜드 셀 수정
    - 중복 옵션 행 삭제
    - 누락 필수 옵션 행 추가
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl 미설치")

    wb = load_workbook(str(src_path), keep_vba=(src_ext == ".xlsm"))
    ws = wb.active

    # ─ 1) 카테고리·브랜드 셀 수정 ─────────────────────────────
    for fr in fix_results:
        cat_changed   = fr.new_category_id and fr.new_category_id != fr.old_category_id
        brand_changed = (not fr.brand_locked) and fr.new_brand and fr.new_brand != fr.old_brand

        if not cat_changed and not brand_changed:
            continue

        for r in fr.row_indices:
            if cat_changed and "category_id" in col_map:
                raw = fr.new_category_id.strip()
                ws.cell(row=r, column=col_map["category_id"],
                        value=int(raw) if raw.isdigit() else raw)
            if cat_changed and fr.new_gosisi_cat and "_gosisi_cat" in col_map:
                ws.cell(row=r, column=col_map["_gosisi_cat"], value=fr.new_gosisi_cat)
            if brand_changed and "brand" in col_map:
                ws.cell(row=r, column=col_map["brand"], value=fr.new_brand)

    # ─ 2) 중복 행 삭제 (큰 인덱스부터) ──────────────────────
    all_dup_rows: list[int] = []
    for fr in fix_results:
        all_dup_rows.extend(fr.dup_rows)

    deleted = 0
    for r in sorted(set(all_dup_rows), reverse=True):
        ws.delete_rows(r)
        deleted += 1
    if deleted:
        print(f"[ExcelFixer] 중복 행 {deleted}개 삭제")

    # ─ 3) 누락 필수 옵션 행 추가 ──────────────────────────────
    added = 0
    for fr in fix_results:
        if not fr.missing_opt_values:
            continue
        # 이 상품의 비중복 첫 번째 행을 템플릿으로 사용
        valid_rows = [r for r in fr.row_indices if r not in fr.dup_rows]
        if not valid_rows:
            continue
        template_row = valid_rows[0]
        max_col = ws.max_column

        for opt_type, opt_val in fr.missing_opt_values.items():
            new_r = ws.max_row + 1
            # 템플릿 행 전체 복사
            for c in range(1, max_col + 1):
                src_cell = ws.cell(template_row, c)
                dst_cell = ws.cell(new_r, c)
                dst_cell.value = src_cell.value
            # 옵션유형1/옵션값1 덮어쓰기
            if "_option_type_0" in col_map:
                ws.cell(new_r, col_map["_option_type_0"]).value = opt_type
            if "_option_val_0" in col_map:
                ws.cell(new_r, col_map["_option_val_0"]).value = opt_val
            # 옵션유형2/3 비움 (이 행은 단일 옵션 차원)
            for i in range(1, 3):
                if f"_option_type_{i}" in col_map:
                    ws.cell(new_r, col_map[f"_option_type_{i}"]).value = None
                if f"_option_val_{i}" in col_map:
                    ws.cell(new_r, col_map[f"_option_val_{i}"]).value = None
            added += 1
            print(f"[ExcelFixer] 필수 옵션 행 추가: {fr.product_name[:25]} | {opt_type}={opt_val}")

    if added:
        print(f"[ExcelFixer] 필수 옵션 행 총 {added}개 추가")

    output_dir.mkdir(parents=True, exist_ok=True)
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"fixed_{ts}{src_ext}"
    wb.save(str(out_path))
    print(f"[ExcelFixer] 수정 저장: {out_path}")
    return out_path
